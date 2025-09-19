import pandas as pd
import os, webbrowser
from datetime import datetime
from PySide6.QtWidgets import QFileDialog, QMessageBox
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook,  load_workbook

class CSV:
    """
    Classe para exportação de dados do sistema de bombeamento para formato XLSX.
    Usa pandas para criar uma planilha formatada com cores e estilo.
    """
    
    def __init__(self):
        self.dados = []
        self.cabecalho = ['Parâmetro', 'Valor', 'Unidade', 'Data/Hora', 'Origem']
        self.df = pd.DataFrame(columns=self.cabecalho)
    
    def adicionar_dados(self, parametro, valor, unidade, origem="Sistema"):
        """
        Adiciona um conjunto de dados para exportação.
        
        Args:
            parametro (str): Nome do parâmetro (ex: "Vazão", "Pressão")
            valor (float/int/str): Valor do parâmetro
            unidade (str): Unidade de medida
            origem (str): Origem dos dados (ex: "Sensor", "Cálculo", "Manual")
        """
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        novo_dado = pd.DataFrame({
            'Parâmetro': [parametro],
            'Valor': [str(valor)],
            'Unidade': [unidade],
            'Data/Hora': [timestamp],
        })
        self.df = pd.concat([self.df, novo_dado], ignore_index=True)
    
    def adicionar_conjunto_dados(self, dados_dict, origem="Sistema"):
        """
        Adiciona um conjunto completo de dados no formato de dicionário.
        
        Args:
            dados_dict (dict): Dicionário no formato {parametro: (valor, unidade)}
            origem (str): Origem dos dados
        """
        for parametro, (valor, unidade) in dados_dict.items():
            self.adicionar_dados(parametro, valor, unidade, origem)
    
    def adicionar_secao(self, nome_secao):
        """
        Adiciona uma linha de seção para organizar o XLSX.
        
        Args:
            nome_secao (str): Nome da seção
        """
        secao_df = pd.DataFrame({
            'Parâmetro': [f"--- {nome_secao} ---"],
            'Valor': [""],
            'Unidade': [""],
            'Data/Hora': [""]

        })
        self.df = pd.concat([self.df, secao_df], ignore_index=True)
    
    def adicionar_espaco(self):
        """Adiciona uma linha em branco para melhor organização."""
        espaco_df = pd.DataFrame({
            'Parâmetro': [""],
            'Valor': [""],
            'Unidade': [""],
            'Data/Hora': [""]
        })
        self.df = pd.concat([self.df, espaco_df], ignore_index=True)
    
    def adicionar_metadados(self, titulo_relatorio, versao_sistema="1.0", info_adicional=""):
        """
        Adiciona metadados ao início do XLSX.
        
        Args:
            titulo_relatorio (str): Título do relatório
            versao_sistema (str): Versão do sistema
            info_adicional (str): Informações adicionais
        """
        metadados = pd.DataFrame({
            'Parâmetro': [
                "Relatório do Sistema de Bombeamento",
                f"Título: {titulo_relatorio}",
                f"Data de geração: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Versão do sistema: {versao_sistema}",
                f"Informações: {info_adicional}",
                ""
            ],
            'Valor': ["", "", "", "", "", ""],
            'Unidade': ["", "", "", "", "", ""],
            'Data/Hora': ["", "", "", "", "", ""],

        })
        
        self.df = pd.concat([metadados, self.df], ignore_index=True)
  
    def aplicar_formato_excel(self, writer, nome_planilha='Relatório'):
        """
        Aplica formatação à planilha Excel usando openpyxl.
        
        Args:
            writer: ExcelWriter do pandas
            nome_planilha (str): Nome da planilha a ser formatada
        """
        # Acessar a planilha e a workbook
        workbook = writer.book
        worksheet = writer.sheets[nome_planilha]
        
        # Definir estilos
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        section_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        
        # Formatar cabeçalho
        for col_num, value in enumerate(self.df.columns.values, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border
            
            # Ajustar largura da coluna
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 20
        
        # Formatar células de dados
        for row in range(2, len(self.df) + 2):
            for col in range(1, len(self.df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Formatar células de seção
                if row <= len(self.df) + 1 and col == 1:
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and str(cell_value).startswith('---'):
                        for sec_col in range(1, len(self.df.columns) + 1):
                            sec_cell = worksheet.cell(row=row, column=sec_col)
                            sec_cell.fill = section_fill
                            sec_cell.font = Font(bold=True)
                            sec_cell.alignment = align_center
            
    def exportar(self, caminho_arquivo=None, parent=None):
        """
        Exporta os dados para um arquivo XLSX com metadados.
        """
        try:
            caminho_arquivo, _ = QFileDialog.getSaveFileName(
                parent, "Salvar como XLSX", "Relatório.xlsx", "Arquivos Excel (*.xlsx)"
            )
            if not caminho_arquivo:
                return False  

            # Garantir extensão .xlsx
            if not caminho_arquivo.lower().endswith('.xlsx'):
                caminho_arquivo += '.xlsx'

            # 1) Escrever dados no arquivo XLSX
            with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
                self.df.to_excel(writer, index=False, sheet_name='Relatório')
                self.aplicar_formato_excel(writer)

            # 2) Reabrir o workbook e adicionar metadados
            wb = load_workbook(caminho_arquivo)

            wb.properties.creator = "AquaPump Software"
            wb.properties.lastModifiedBy = "AquaPump Software"
            wb.properties.title = "Relatório de Bombas"
            wb.properties.subject = "Dimensionamento hidráulico"
            wb.properties.description = "Planilha gerada automaticamente com openpyxl"
            wb.properties.keywords = "hidráulica, bombas, engenharia"
            wb.properties.category = "Projetos"
            wb.properties.created = datetime.now()
            wb.properties.modified = datetime.now()
            wb.properties.version = "1.0"
            wb.properties.company = "AquaPump"
            wb.properties.manager = "Tenerife Saugineta"
            wb.properties.comments = "Gerado por AquaPump Software"

            wb.save(caminho_arquivo)

            # 4) Abrir no sistema
            caminho_absoluto = os.path.abspath(caminho_arquivo)
            webbrowser.open_new(f"file://{caminho_absoluto}")
            return True

        except Exception as e:
            print("Erro ao exportar:", e)
            return False

    
    def limpar_dados(self):
        """Limpa todos os dados armazenados para uma nova exportação."""
        self.df = pd.DataFrame(columns=self.cabecalho)
    
    def gerar_relatorio_completo(self, dados_bombeamento, dados_analise, dados_ambiente, caminho_arquivo=None, parent=None):
        """
        Gera um relatório XLSX completo com todas as seções organizadas.
        
        Args:
            dados_bombeamento (dict): Dados do sistema de bombeamento
            dados_analise (dict): Dados de análise e cálculos
            dados_ambiente (dict): Dados ambientais
            caminho_arquivo (str): Caminho para salvar o arquivo
            parent: Widget pai para os diálogos
        """
        self.limpar_dados()
        
        # Adicionar metadados
        self.adicionar_metadados(
            "Relatório Completo do Sistema de Bombeamento",
            "1.0",
            "Relatório gerado automaticamente pelo sistema"
        )
        
        # Seção de dados do bombeamento
        self.adicionar_secao("DADOS DO SISTEMA DE BOMBEAMENTO")
        self.adicionar_conjunto_dados(dados_bombeamento)
        self.adicionar_espaco()
        
        # Seção de análise
        self.adicionar_secao("ANÁLISE E CÁLCULOS")
        self.adicionar_conjunto_dados(dados_analise, "Análise do Sistema")
        self.adicionar_espaco()
        
        # Seção ambiental
        self.adicionar_secao("CONDIÇÕES AMBIENTAIS")
        self.adicionar_conjunto_dados(dados_ambiente, "Sensores Ambientais")
        
        # Exportar
        return self.exportar(caminho_arquivo, parent)